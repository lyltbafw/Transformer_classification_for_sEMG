import torch
import torch.nn as nn
import numpy as np
import openpyxl
import paramiko

def ssh_connect(hostname, username, password):
    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    client.connect(hostname, username=username, password=password)
    return client
def excel2m(path):
    #print(path)
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    nrows = sheet.max_row - 1  # 行数，减去第一行
    ncols_to_extract = [3, 4, 7, 8]  # 需要提取的列
    datamatrix = np.matrix(np.zeros((nrows, len(ncols_to_extract))))  # 使用 numpy.matrix

    for idx, col_num in enumerate(ncols_to_extract):
        cols = [sheet.cell(row=row, column=col_num).value for row in range(2, nrows + 2)]
        # 过滤掉非数值类型的数据
        cols = [c for c in cols if isinstance(c, (int, float))]
        if not cols:
            continue  # 如果列中没有数值类型数据，则跳过
        minVals = min(cols)
        maxVals = max(cols)
        cols1 = np.matrix(cols).transpose()  # 把list转换为numpy矩阵进行操作
        ranges = maxVals - minVals
        if ranges == 0:
            ranges = 1  # 防止除以零
        b = cols1 - minVals
        normcols = b / ranges  # 数据进行归一化处理
        # 将归一化后的数据存储到 datamatrix 的相应行
        for i in range(len(normcols)):
            datamatrix[i, idx] = normcols[i, 0]

    return datamatrix.A  # 转置，使得每一行代表一个特征

# 定义PositionalEncoding
class PositionalEncoding(nn.Module):
    def __init__(self, d_model, dropout, max_len=5000):
        super(PositionalEncoding, self).__init__()
        self.dropout = nn.Dropout(p=dropout)

        # pe的维度(位置编码最大长度，模型维度)
        pe = torch.zeros(max_len, d_model)
        # 维度为（max_len, 1）：先在[0,max_len]中取max_len个整数，再加一个维度
        position = torch.arange(0, max_len, dtype=torch.float).unsqueeze(1)
        # 位置编码的除数项：10000^(2i/d_model)
        div_term = torch.exp(torch.arange(0, d_model, 2).float() * (-np.log(10000.0) / d_model))
        # sin负责奇数；cos负责偶数
        pe[:, 0::2] = torch.sin(position * div_term)
        pe[:, 1::2] = torch.cos(position * div_term)

        # 维度变换：(max_len,d_model)→(1,max_len,d_model)→(max_len,1,d_model)
        pe = pe.unsqueeze(0).transpose(0, 1)
        # 将pe注册为模型缓冲区
        self.register_buffer('pe', pe)

    def forward(self, x):
        #print(x.shape)
        #print(self.pe[:x.size(0), :].shape)
        # 取pe的前x.size(0)行，即
        # (x.size(0),1,d_model) → (x.size(0),d_model)，拼接到x上
        x = x + self.pe[:x.size(0), :]
        return self.dropout(x)


# 定义Transformer模型
class TransformerModel(nn.Module):
    def __init__(self, input_dim, output_dim, d_model, nhead, num_layers, dim_feedforward, dropout):
        super(TransformerModel, self).__init__()
        # 创建一个线性变换层，维度input_dim4→d_model
        self.embedding = nn.Linear(input_dim, d_model)  # 使用嵌入层
        # 生成pe
        self.pos_encoder = PositionalEncoding(d_model, dropout)
        # 生成一层encoder
        encoder_layers = nn.TransformerEncoderLayer(d_model=d_model, nhead=nhead, dim_feedforward=dim_feedforward,
                                                    dropout=dropout)
        # 多层encoder
        self.transformer_encoder = nn.TransformerEncoder(encoder_layers, num_layers=num_layers)
        # 维度d_model→output_dim
        self.d_model = d_model
        self.pooling = nn.AdaptiveAvgPool1d(1)

        # 分类头
        self.classifier = nn.Linear(d_model, output_dim)
    def forward(self, src):
        # 嵌入输入
        src = self.embedding(src.float())
        #print(src.shape)
        # 加上位置嵌入
        src = self.pos_encoder(src)
        src = src.permute(1, 0, 2)  # [sequence_length, batch_size, d_model]，因为Transformer期望输入的第一个维度是序列长度
        output = self.transformer_encoder(src)
        # 调整输出形状为(batch, seq_len, d_model)
        output = output.permute(1, 0, 2)
        # 对所有位置的表示取平均
        output = output.permute(0, 2, 1)  # [batch_size, d_model, sequence_length]
        output = self.pooling(output)  # [batch_size, d_model, 1]
        output = output.squeeze(-1)  # [batch_size, d_model]
        # 线性变换
        output = self.classifier(output)
        # 使用sigmoid激活函数
        return output


def predict(model, data):
    model.eval()
    with torch.no_grad():
        data = torch.from_numpy(data).float().to(device)
        data = data.unsqueeze(0)
        output = model(data)
        output = output.softmax(dim=1)
    return output.cpu().numpy()

# 开始训
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
test_path = "Datasets/arm/test/12-1.xlsx"
data = excel2m(test_path)
data = data[0:1500,:]
# 初始化模型
model = TransformerModel(input_dim=4,
                         output_dim=3,
                         d_model=128,
                         nhead=8,
                         num_layers=3,
                         dim_feedforward=256,
                         dropout=0.1).to(device)
model.load_state_dict(torch.load("models/transformer_model_Adam6.pth.pth"))
output = predict(model, data)
reuslt = output.argmax()
order = ['arm-up', 'grasp', 'hand-up']
print(order[reuslt])
hostname = '192.168.175.164'
username = 'nao'
password = 'nao'
client = ssh_connect(hostname, username, password)
print(client)
if reuslt == 0:
    stdin, stdout, stderr = client.exec_command('python pycall_NaoBendarm.py')
elif reuslt == 1:
    stdin, stdout, stderr = client.exec_command('python pycall_NaoGrasp.py')
else:
    stdin, stdout, stderr = client.exec_command('python pycall_NaoHandsup.py')
print(stdout.read().decode())
client.close()

