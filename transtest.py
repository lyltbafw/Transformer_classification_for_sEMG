import openpyxl
import numpy as np


def excel2m(path):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    nrows = sheet.max_row - 1  # 行数，减去第一行
    ncols_to_extract = [3, 4, 7, 8]  # 需要提取的列
    datamatrix = np.matrix(np.zeros((nrows, len(ncols_to_extract))))  # 使用 numpy.matrix

    for idx, col_num in enumerate(ncols_to_extract):
        cols = [sheet.cell(row=row, column=col_num).value for row in range(2, nrows + 2)]
        # 过滤掉非数值类型的数据
        cols = [c for c in cols if isinstance(c, (int, float))]
        if not cols:
            continue  # 如果列中没有数值类型数据，则跳过
        minVals = min(cols)
        maxVals = max(cols)
        cols1 = np.matrix(cols).transpose()  # 把list转换为numpy矩阵进行操作
        ranges = maxVals - minVals
        if ranges == 0:
            ranges = 1  # 防止除以零
        b = cols1 - minVals
        normcols = b / ranges  # 数据进行归一化处理
        # 将归一化后的数据存储到 datamatrix 的相应行
        for i in range(len(normcols)):
            datamatrix[i, idx] = normcols[i, 0]

    return datamatrix


path1 = r"D:\learn\pythonProject\trans\data\bendarm\1.xlsx"
data_matrix = excel2m(path1)
print(data_matrix)
