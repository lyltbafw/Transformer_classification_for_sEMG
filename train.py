import torch
import torch.nn as nn
import torch.optim as optim
import numpy as np
import pandas as pd
import openpyxl
import os
from torch.utils.data import Dataset
from torch.nn.utils.rnn import pad_sequence
import matplotlib.pyplot as plt
# 初始化损失记录
train_losses = []
val_losses = []

# 绘制损失曲线的函数
def plot_loss_curve(train_losses, val_losses):
    epochs = range(1, len(train_losses) + 1)

    plt.figure(figsize=(10, 6))
    plt.plot(epochs, train_losses, 'b', label='Training Loss')
    plt.plot(epochs, val_losses, 'r', label='Validation Loss')
    plt.title('Training and Validation Loss')
    plt.xlabel('Epochs')
    plt.ylabel('Loss')
    plt.legend()
    plt.grid(True)
    plt.show()
def excel2m(path):
    #print(path)
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    nrows = sheet.max_row - 1  # 行数，减去第一行
    ncols_to_extract = [3, 4, 7, 8]  # 需要提取的列
    datamatrix = np.matrix(np.zeros((nrows, len(ncols_to_extract))))  # 使用 numpy.matrix

    for idx, col_num in enumerate(ncols_to_extract):
        cols = [sheet.cell(row=row, column=col_num).value for row in range(2, nrows + 2)]
        # 过滤掉非数值类型的数据
        cols = [c for c in cols if isinstance(c, (int, float))]
        if not cols:
            continue  # 如果列中没有数值类型数据，则跳过
        minVals = min(cols)
        maxVals = max(cols)
        cols1 = np.matrix(cols).transpose()  # 把list转换为numpy矩阵进行操作
        ranges = maxVals - minVals
        if ranges == 0:
            ranges = 1  # 防止除以零
        b = cols1 - minVals
        normcols = b / ranges  # 数据进行归一化处理
        # 将归一化后的数据存储到 datamatrix 的相应行
        for i in range(len(normcols)):
            datamatrix[i, idx] = normcols[i, 0]

    return datamatrix.A  # 转置，使得每一行代表一个特征

class myDataset(Dataset):
    def __init__(self, data_dir, transform=None):
        """
        初始化数据集
        :param data_dir: 包含数据的目录路径
        :param transform: 可选的数据转换函数/变换
        """
        self.data_dir = data_dir
        self.transform = transform
        # 获取所有文件名
        self.file_names = [f for f in os.listdir(data_dir) if os.path.isfile(os.path.join(data_dir, f))]
        # 确保文件名以数字结尾（即标签）
        self.file_names = [f for f in self.file_names if f[-6].isdigit()]  # 假设文件扩展名为4个字符长，如'.txt'

    def __len__(self):
        """
        返回数据集的大小
        """
        return len(self.file_names)

    def __getitem__(self, idx):
        """
        根据索引获取数据项
        :param idx: 数据项的索引
        """
        if idx >= len(self):
            raise IndexError("Index out of range")

        # 获取文件名
        file_name = self.file_names[idx]
        # 文件路径
        file_path = os.path.join(self.data_dir, file_name)
        # 读取文件内容
        data = excel2m(file_path)
        data = data[0:1500,:]
        #print(data.shape)
        # 获取标签（假设标签是文件名的最后一个字符）
        label = torch.zeros(3)
        label[int(file_name[-6])-1] = 1
        #label = int(file_name[-6])  # 假设文件扩展名为4个字符长，如'.txt'
        #print(label.shape)

        # 如果有transform操作，则应用到数据上
        if self.transform:
            data = self.transform(data)
        # 返回数据和标签
        return data, label

# 定义PositionalEncoding
class PositionalEncoding(nn.Module):
    def __init__(self, d_model, dropout, max_len=5000):
        super(PositionalEncoding, self).__init__()
        self.dropout = nn.Dropout(p=dropout)

        # pe的维度(位置编码最大长度，模型维度)
        pe = torch.zeros(max_len, d_model)
        # 维度为（max_len, 1）：先在[0,max_len]中取max_len个整数，再加一个维度
        position = torch.arange(0, max_len, dtype=torch.float).unsqueeze(1)
        # 位置编码的除数项：10000^(2i/d_model)
        div_term = torch.exp(torch.arange(0, d_model, 2).float() * (-np.log(10000.0) / d_model))
        # sin负责奇数；cos负责偶数
        pe[:, 0::2] = torch.sin(position * div_term)
        pe[:, 1::2] = torch.cos(position * div_term)

        # 维度变换：(max_len,d_model)→(1,max_len,d_model)→(max_len,1,d_model)
        pe = pe.unsqueeze(0).transpose(0, 1)
        # 将pe注册为模型缓冲区
        self.register_buffer('pe', pe)

    def forward(self, x):
        #print(x.shape)
        #print(self.pe[:x.size(0), :].shape)
        # 取pe的前x.size(0)行，即
        # (x.size(0),1,d_model) → (x.size(0),d_model)，拼接到x上
        x = x + self.pe[:x.size(0), :]
        return self.dropout(x)


# 定义Transformer模型
class TransformerModel(nn.Module):
    def __init__(self, input_dim, output_dim, d_model, nhead, num_layers, dim_feedforward, dropout):
        super(TransformerModel, self).__init__()
        # 创建一个线性变换层，维度input_dim4→d_model
        self.embedding = nn.Linear(input_dim, d_model)  # 使用嵌入层
        # 生成pe
        self.pos_encoder = PositionalEncoding(d_model, dropout)
        # 生成一层encoder
        encoder_layers = nn.TransformerEncoderLayer(d_model=d_model, nhead=nhead, dim_feedforward=dim_feedforward,
                                                    dropout=dropout)
        # 多层encoder
        self.transformer_encoder = nn.TransformerEncoder(encoder_layers, num_layers=num_layers)
        # 维度d_model→output_dim
        self.d_model = d_model
        self.pooling = nn.AdaptiveAvgPool1d(1)

        # 分类头
        self.classifier = nn.Linear(d_model, output_dim)
    def forward(self, src):
        # 嵌入输入
        src = self.embedding(src.float())
        #print(src.shape)
        # 加上位置嵌入
        src = self.pos_encoder(src)
        src = src.permute(1, 0, 2)  # [sequence_length, batch_size, d_model]，因为Transformer期望输入的第一个维度是序列长度
        output = self.transformer_encoder(src)
        # 调整输出形状为(batch, seq_len, d_model)
        output = output.permute(1, 0, 2)
        # 对所有位置的表示取平均
        output = output.permute(0, 2, 1)  # [batch_size, d_model, sequence_length]
        output = self.pooling(output)  # [batch_size, d_model, 1]
        output = output.squeeze(-1)  # [batch_size, d_model]
        # 线性变换
        output = self.classifier(output)
        # 使用sigmoid激活函数
        return output

# 训练模型
def train(model, iterator, optimizer, criterion):
    # iterator是train_loader= DataLoader(train_dataset, batch_size=32)
    model.train()
    running_loss = 0.0
    for batch in iterator:
        # 初始化，防止梯度爆炸
        optimizer.zero_grad()

        X, y = batch
        X = X.to(device)
        y = y.to(device)
        predictions = model(X)

        # 计算修正损失函数
        loss = criterion(predictions.squeeze(), y)
        # 计算当前批数据的损失函数对模型参数的梯度
        loss.backward()
        # 根据梯度更新模型参数
        optimizer.step()
        # 累加损失
        running_loss += loss.item()
    epoch_train_loss = running_loss / len(train_loader)
    train_losses.append(epoch_train_loss)
# 测试模型
def evaluate(model, iterator, criterion):
    print("===========test==========")
    model.eval()
    epoch_loss = 0

    with torch.no_grad():  # 此处不需要梯度计算
        for batch in iterator:
            X, y = batch
            X = X.to(device)
            y = y.to(device)
            predictions = model(X)
            loss = criterion(predictions.squeeze(), y)
            epoch_loss += loss.item()
        # 计算平均损失
        epoch_loss = epoch_loss / len(iterator)
        val_losses.append(epoch_loss)
    # 累加loss，再求平均
    return epoch_loss


# 开始训
N_EPOCHS = 300
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
train_path = "Datasets/arm/train"
test_path = "Datasets/arm/test"

train_dataset = myDataset(train_path)
test_dataset = myDataset(test_path)

train_loader = torch.utils.data.DataLoader(train_dataset, batch_size=4, shuffle=True, num_workers=4)
test_loader = torch.utils.data.DataLoader(test_dataset, batch_size=3, shuffle=False, num_workers=4)
# 初始化模型
model = TransformerModel(input_dim=4,
                         output_dim=3,
                         d_model=128,
                         nhead=8,
                         num_layers=3,
                         dim_feedforward=256,
                         dropout=0.1).to(device)
# 定义损失函数和优化器
criterion = nn.CrossEntropyLoss(weight=torch.FloatTensor([1, 1, 1]).to(device))
optimizer = optim.Adam(model.parameters(), lr=0.00003, betas=(0.9, 0.999), eps=1e-08, weight_decay=0, amsgrad=False)

for epoch in range(N_EPOCHS):
    print("第%d轮====================================" % (epoch + 1))
    train(model, train_loader, optimizer, criterion)
    test_loss = evaluate(model, test_loader, criterion)
    print(f'Epoch: {epoch + 1:02}, Test Loss: {test_loss:.3f}')
    #torch.save(model.state_dict(), 'models/transformer_model_Adam.pth')
# 保存模型
torch.save(model.state_dict(), 'models/transformer_model_Adam6.pth.pth')
# 绘制损失曲线
plot_loss_curve(train_losses, val_losses)